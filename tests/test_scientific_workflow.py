"""Integration Tests für Scientific Methodology Workflow"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from models import (
    Config, CheckAttribute, ScientificConfig, AnalysisResult,
    ProcessingStats, MergedResult
)
from config_manager import ConfigManager
from confidence_engine import ConfidenceEngine
from kappa_calculator import KappaCalculator
from multi_coder import MultiCoder
from reproducibility_manager import ReproducibilityManager, MethodologyMetadata
from output_manager import OutputManager
from excel_writer import ExcelWriter
from statistics_generator import StatisticsGenerator
import tempfile
import json


class MockAnalyzer:
    """Mock LLMAnalyzer für Integration Tests"""
    
    def __init__(self, sentiment="positiv", check_value=True, confidence=85):
        self.sentiment = sentiment
        self.check_value = check_value
        self.confidence = confidence
    
    def analyze_text(self, text, check_attributes, **kwargs):
        result = AnalysisResult(
            paraphrase=f"Paraphrase von: {text[:30]}...",
            sentiment=self.sentiment,
            sentiment_reason=f"Test Grund fuer {text[:20]}",
            keywords=["test", "wort", "analyse"],
            custom_checks={attr.question: self.check_value for attr in check_attributes},
            custom_checks_reasons={attr.question: "Testgrund" for attr in check_attributes}
        )
        
        # Füge Confidence hinzu
        for attr in check_attributes:
            result.add_confidence(
                attr.question, 
                self.confidence, 
                "Sicher"
            )
        
        return result


class TestScientificWorkflow:
    """Integration Tests für den gesamten wissenschaftlichen Workflow"""
    
    def test_end_to_end_excel_mode(self):
        """End-to-End Test für Excel-Modus mit wissenschaftlichen Parametern"""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)
            
            # 1. Config erstellen
            attr = CheckAttribute(
                question="Engagement gefoerdert?",
                answer_type="boolean",
                definition="Prueft ob Engagement gefoerdert wird"
            )
            attr2 = CheckAttribute(
                question="Kategorie?",
                answer_type="categorical",
                categories=["Finanzierung", "Bildung", "Gesundheit"]
            )
            
            sc = ScientificConfig(
                multi_coder=True,
                coder_models=["gpt-4o-mini", "gpt-4o"],
                primary_coder="highest_confidence",
                confidence_threshold=70,
                seed=42
            )
            
            config = Config(
                check_attributes=[attr, attr2],
                model="gpt-4o-mini",
                provider="openai",
                research_question="Wie beeinflusst Engagement die Gesellschaft?",
                scientific=sc
            )
            
            # 2. Multi-Coder initialisieren
            mc = MultiCoder(sc)
            mc.add_analyzer("gpt-4o-mini", MockAnalyzer("positiv", True, 85))
            mc.add_analyzer("gpt-4o", MockAnalyzer("positiv", True, 90))
            
            # 3. Texte analysieren
            texts = [
                "Das Programm foerdert Engagement in der Bildung.",
                "Finanzielle Unterstuetzung ist wichtig.",
                "Gesundheitsprojekte werden unterstuetzt."
            ]
            
            all_results = []
            intercoder_results = []
            
            for text in texts:
                # Einzelanalyse
                analyzer = MockAnalyzer("positiv", True, 85)
                result = analyzer.analyze_text(text, [attr, attr2])
                all_results.append(result)
                
                # Multi-Coder Analyse
                ic_result = mc.encode_text(text, [attr, attr2])
                intercoder_results.append(ic_result)
            
            # 4. Confidence Engine
            ce = ConfidenceEngine()
            for result in all_results:
                for a in [attr, attr2]:
                    score = result.confidence_scores.get(a.question, 0)
                    if score < 0.7:
                        result.low_confidence_flags.append(a.question)
            
            # 5. Kappa berechnen
            kappa_results = {}
            for ic in intercoder_results:
                for question in ic.agreements:
                    if question not in kappa_results:
                        kappa_results[question] = []
                    kappa_results[question].append(1.0 if ic.agreements[question] else 0.0)
            
            # 6. Output erstellen
            output_manager = OutputManager(tmpdir)
            
            # Excel Writer
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            ew = ExcelWriter(confidence_threshold=70)
            ew.create_intercoder_sheet(wb, intercoder_results[0], [attr, attr2])
            ew.create_kappa_sheet(wb, intercoder_results[0])
            
            excel_path = output_manager.get_analyzed_path("test")
            wb.save(excel_path)
            
            # 7. Reproduzierbarkeit
            rm = ReproducibilityManager(output_manager.dirs["reproducibility"])
            
            metadata = MethodologyMetadata(
                model=config.model,
                provider=config.provider,
                temperature=0.3,
                seed=config.scientific.seed,
                research_question=config.research_question,
                check_attributes=[
                    {"question": a.question, "answer_type": a.answer_type,
                     "categories": a.categories or [], "definition": a.definition or ""}
                    for a in config.check_attributes
                ],
                multi_coder=True,
                coder_models=sc.coder_models,
                confidence_threshold=sc.confidence_threshold,
                total_items=len(texts),
                total_successful=len(texts),
                total_failed=0,
                total_tokens=15000
            )
            
            results_dicts = [
                {"custom_checks": r.custom_checks, "sentiment": r.sentiment, "keywords": r.keywords}
                for r in all_results
            ]
            
            output_files = rm.save_all(
                metadata,
                [{"question": a.question, "answer_type": a.answer_type,
                  "categories": a.categories or [], "definition": a.definition or ""}
                 for a in config.check_attributes],
                results_dicts
            )
            
            # 8. Validierung
            assert excel_path.exists()
            assert all(p.exists() for p in output_files.values())
            assert len(all_results) == 3
            assert len(intercoder_results) == 3
            
            # Prüfe Confidence Scores
            for result in all_results:
                assert len(result.confidence_scores) == 2
            
            # Prüfe Kappa
            for question in kappa_results:
                avg_kappa = sum(kappa_results[question]) / len(kappa_results[question])
                assert avg_kappa >= 0
            
            print("  OK End-to-End Excel-Modus OK")
    
    def test_confidence_workflow(self):
        """Test des Confidence-Workflows"""
        # 1. Confidence aus LLM-Antwort extrahieren
        response = {
            "confidence": {"Engagement?": 85, "Kategorie?": 45},
            "confidence_reasons": {"Engagement?": "Sicher", "Kategorie?": "Unsicher"},
            "alternatives": {"Kategorie?": ["Finanzierung", "Bildung"]}
        }
        
        attrs = [
            {"question": "Engagement?", "answer_type": "boolean"},
            {"question": "Kategorie?", "answer_type": "categorical"}
        ]
        
        results = ConfidenceEngine.extract_confidence_from_response(response, attrs, threshold=0.7)
        
        # 2. Prüfe Ergebnisse
        assert results["Engagement?"]["score"] == 0.85
        assert results["Engagement?"]["is_low"] == False
        assert results["Kategorie?"]["score"] == 0.45
        assert results["Kategorie?"]["is_low"] == True
        
        # 3. Statistiken
        stats = ConfidenceEngine.get_summary_statistics(results)
        assert stats["count"] == 2
        assert stats["low_count"] == 1
        
        print("  OK Confidence Workflow OK")
    
    def test_intercoder_workflow(self):
        """Test des Intercoder-Workflows"""
        # 1. MultiCoder initialisieren
        sc = ScientificConfig(
            multi_coder=True,
            coder_models=["model1", "model2"],
            primary_coder="highest_confidence"
        )
        
        mc = MultiCoder(sc)
        mc.add_analyzer("model1", MockAnalyzer("positiv", True, 90))
        mc.add_analyzer("model2", MockAnalyzer("positiv", True, 85))
        
        attrs = [
            CheckAttribute(question="Q1?", answer_type="boolean"),
            CheckAttribute(question="Q2?", answer_type="categorical", categories=["A", "B"])
        ]
        
        # 2. Kodierung
        result = mc.encode_text("Testtext", attrs)
        
        # 3. Validierung
        assert len(result.coder_results) == 2
        assert result.primary_coder.model_name in ["model1", "model2"]
        
        # 4. Kappa berechnen
        kappa, ci, interp = KappaCalculator.cohens_kappa(
            ["True", "A"],
            ["True", "A"]
        )
        assert kappa == 1.0
        
        print("  OK Intercoder Workflow OK")
    
    def test_reproducibility_workflow(self):
        """Test des Reproduzierbarkeits-Workflows"""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)
            
            # 1. ReproducibilityManager initialisieren
            rm = ReproducibilityManager(tmpdir)
            
            # 2. Audit-Einträge aufzeichnen
            for i in range(5):
                rm.record_analysis(
                    model="gpt-4o-mini",
                    provider="openai",
                    prompt=f"Prompt {i}",
                    response=f"Response {i}",
                    seed=42,
                    input_text=f"Input {i}"
                )
            
            # 3. Methodology generieren
            metadata = MethodologyMetadata(
                model="gpt-4o-mini",
                provider="openai",
                temperature=0.3,
                seed=42,
                research_question="Test?",
                check_attributes=[
                    {"question": "Q1?", "answer_type": "boolean"}
                ],
                total_items=5,
                total_successful=5,
                total_failed=0,
                total_tokens=25000
            )
            
            output = rm.save_all(
                metadata,
                [{"question": "Q1?", "answer_type": "boolean"}]
            )
            
            # 4. Validierung
            assert output["methodology"].exists()
            assert output["codebook"].exists()
            assert output["audit_trail"].exists()
            
            # Prüfe Audit Trail
            with open(output["audit_trail"], 'r', encoding='utf-8') as f:
                audit_data = json.load(f)
            assert audit_data["metadata"]["total_entries"] == 5
            
            # Prüfe Methodology
            meth_content = output["methodology"].read_text(encoding='utf-8')
            assert "gpt-4o-mini" in meth_content
            assert "Test?" in meth_content
            
            print("  OK Reproducibility Workflow OK")
    
    def test_excel_output_workflow(self):
        """Test des Excel-Output-Workflows"""
        from openpyxl import Workbook
        
        # 1. Mock-Daten erstellen
        attrs = [
            CheckAttribute(question="Q1?", answer_type="boolean"),
            CheckAttribute(question="Q2?", answer_type="categorical", categories=["A", "B"])
        ]
        
        results = []
        for i in range(5):
            result = AnalysisResult(
                paraphrase=f"Paraphrase {i}",
                sentiment="positiv",
                sentiment_reason="Grund",
                keywords=["test", "wort"],
                custom_checks={"Q1?": i % 2 == 0, "Q2?": "A" if i % 2 == 0 else "B"}
            )
            result.add_confidence("Q1?", 85, "Sicher")
            result.add_confidence("Q2?", 45 if i % 2 == 0 else 80, "Test")
            results.append(result)
        
        # 2. Excel Writer testen
        ew = ExcelWriter(confidence_threshold=70)
        wb = Workbook()
        wb.remove(wb.active)
        
        # Header testen
        headers = ew._build_result_headers(attrs, include_reasoning=True, include_confidence=True)
        assert "Q1? (Konfidenz)" in headers
        assert "Q2? (Konfidenz)" in headers
        
        # Confidence-Schreiben testen
        from openpyxl.worksheet.worksheet import Worksheet
        ws = wb.create_sheet("Test")
        offset = ew._write_confidence_to_row(ws, 2, 1, results[0], attrs)
        assert offset == 2
        
        print("  OK Excel Output Workflow OK")


def run_tests():
    """Führt alle Integration Tests aus"""
    test = TestScientificWorkflow()
    tests = [
        test.test_end_to_end_excel_mode,
        test.test_confidence_workflow,
        test.test_intercoder_workflow,
        test.test_reproducibility_workflow,
        test.test_excel_output_workflow,
    ]
    
    passed = 0
    failed = 0
    
    for t in tests:
        try:
            t()
            passed += 1
        except AssertionError as e:
            failed += 1
            print(f"  FAIL {t.__name__}: {e}")
        except Exception as e:
            failed += 1
            print(f"  FAIL {t.__name__}: {type(e).__name__}: {e}")
    
    return passed, failed


if __name__ == "__main__":
    print("=== test_scientific_workflow.py ===")
    passed, failed = run_tests()
    print(f"\n{passed} bestanden, {failed} fehlgeschlagen")
