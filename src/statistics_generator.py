"""Statistics Generator für Kategorie-Auswertung"""

from pathlib import Path
from typing import Dict, List, Set
from openpyxl import Workbook
from logging_config import get_logger

logger = get_logger("statistics_generator")


class StatisticsGenerator:
    """Erstellt Auswertungsdatei mit Kategorie-Häufigkeiten"""
    
    def __init__(self):
        """Initialisiert StatisticsGenerator"""
        pass
    
    def calculate_category_frequencies(self, 
                                      category_assignments: List[List[str]]) -> Dict[str, int]:
        """
        Berechnet Häufigkeit jeder Kategorie.
        
        Args:
            category_assignments: Liste von Kategorie-Listen pro Antwort
            
        Returns:
            Dictionary mit Kategorie-Häufigkeiten
        """
        frequencies = {}
        
        for categories in category_assignments:
            for category in categories:
                frequencies[category] = frequencies.get(category, 0) + 1
        
        logger.info(f"Häufigkeiten für {len(frequencies)} Kategorien berechnet")
        return frequencies
    
    def calculate_per_sheet_frequencies(self,
                                       sheet_names: List[str],
                                       sheet_row_counts: List[int],
                                       category_assignments: List[List[str]]) -> Dict[str, Dict[str, int]]:
        """
        Berechnet Häufigkeit jeder Kategorie pro Sheet.
        
        Args:
            sheet_names: Namen der Sheets
            sheet_row_counts: Anzahl Zeilen pro Sheet
            category_assignments: Liste von Kategorie-Listen pro Antwort
            
        Returns:
            Dictionary mit Sheet-Namen als Keys und Kategorie-Häufigkeiten als Values
        """
        per_sheet_frequencies = {}
        
        # Initialisiere für jedes Sheet
        for sheet_name in sheet_names:
            per_sheet_frequencies[sheet_name] = {}
        
        # Verarbeite Ergebnisse pro Sheet
        current_index = 0
        for sheet_name, row_count in zip(sheet_names, sheet_row_counts):
            # Hole die Kategorie-Zuordnungen für dieses Sheet
            sheet_assignments = category_assignments[current_index:current_index + row_count]
            
            # Zähle Kategorien für dieses Sheet
            for categories in sheet_assignments:
                for category in categories:
                    per_sheet_frequencies[sheet_name][category] = \
                        per_sheet_frequencies[sheet_name].get(category, 0) + 1
            
            current_index += row_count
        
        logger.info(f"Per-Sheet-Häufigkeiten für {len(sheet_names)} Sheets berechnet")
        return per_sheet_frequencies
    
    def collect_keywords_per_category(self,
                                     keyword_to_category: Dict[str, str]) -> Dict[str, Set[str]]:
        """
        Sammelt alle Keywords pro Kategorie (invertiert das Mapping).
        
        Args:
            keyword_to_category: Mapping von Keyword -> Kategorie
            
        Returns:
            Dictionary mit Kategorien und deduplizierten Keywords
        """
        keywords_per_category = {}
        
        for keyword, category in keyword_to_category.items():
            if category not in keywords_per_category:
                keywords_per_category[category] = set()
            keywords_per_category[category].add(keyword)
        
        logger.info(f"Keywords für {len(keywords_per_category)} Kategorien gesammelt")
        return keywords_per_category
    
    def create_statistics_workbook(self, 
                                  per_sheet_frequencies: Dict[str, Dict[str, int]],
                                  total_frequencies: Dict[str, int],
                                  keywords_per_category: Dict[str, Set[str]]) -> Workbook:
        """
        Erstellt neue Excel-Datei mit Statistiken (per Sheet + Gesamt).
        
        Args:
            per_sheet_frequencies: Kategorie-Häufigkeiten pro Sheet
            total_frequencies: Gesamt-Kategorie-Häufigkeiten
            keywords_per_category: Keywords pro Kategorie
            
        Returns:
            Workbook mit Statistiken
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Kategorie-Statistiken"
        
        current_row = 1
        
        # Für jedes Sheet eine Sektion erstellen
        for sheet_name, frequencies in per_sheet_frequencies.items():
            # Sheet-Überschrift
            sheet.cell(row=current_row, column=1, value=f"Sheet: {sheet_name}")
            current_row += 1
            
            # Spaltenüberschriften
            sheet.cell(row=current_row, column=1, value="Kategorie")
            sheet.cell(row=current_row, column=2, value="Häufigkeit")
            sheet.cell(row=current_row, column=3, value="Keywords")
            current_row += 1
            
            # Sortiere Kategorien nach Häufigkeit (absteigend)
            sorted_categories = sorted(frequencies.items(), key=lambda x: x[1], reverse=True)
            
            # Schreibe Daten
            for category, frequency in sorted_categories:
                sheet.cell(row=current_row, column=1, value=category)
                sheet.cell(row=current_row, column=2, value=frequency)
                
                # Keywords kommagetrennt
                if category in keywords_per_category:
                    keywords_str = ", ".join(sorted(keywords_per_category[category]))
                    sheet.cell(row=current_row, column=3, value=keywords_str)
                
                current_row += 1
            
            # Leerzeile nach jedem Sheet
            current_row += 1
        
        # "Zusammen" Sektion
        sheet.cell(row=current_row, column=1, value="Zusammen")
        current_row += 1
        
        # Spaltenüberschriften
        sheet.cell(row=current_row, column=1, value="Kategorie")
        sheet.cell(row=current_row, column=2, value="Häufigkeit")
        sheet.cell(row=current_row, column=3, value="Keywords")
        current_row += 1
        
        # Sortiere Gesamt-Kategorien nach Häufigkeit (absteigend)
        sorted_total = sorted(total_frequencies.items(), key=lambda x: x[1], reverse=True)
        
        # Schreibe Gesamt-Daten
        for category, frequency in sorted_total:
            sheet.cell(row=current_row, column=1, value=category)
            sheet.cell(row=current_row, column=2, value=frequency)
            
            # Keywords kommagetrennt
            if category in keywords_per_category:
                keywords_str = ", ".join(sorted(keywords_per_category[category]))
                sheet.cell(row=current_row, column=3, value=keywords_str)
            
            current_row += 1
        
        logger.info(f"Statistik-Workbook mit {len(per_sheet_frequencies)} Sheets und Gesamt-Statistik erstellt")
        return workbook
    
    def save_statistics(self, workbook: Workbook, output_path: Path) -> None:
        """
        Speichert Statistik-Datei.
        
        Args:
            workbook: Zu speicherndes Workbook
            output_path: Zielpfad
        """
        workbook.save(output_path)
        logger.info(f"Statistik-Datei gespeichert: {output_path}")
        print(f"\n✓ Statistik-Datei gespeichert: {output_path}")
    
    def generate_statistics(self,
                           sheet_names: List[str],
                           sheet_row_counts: List[int],
                           category_assignments: List[List[str]],
                           keyword_to_category: Dict[str, str],
                           output_path: Path) -> None:
        """
        Führt vollständige Statistik-Generierung durch.
        
        Args:
            sheet_names: Namen der Sheets
            sheet_row_counts: Anzahl Zeilen pro Sheet
            category_assignments: Kategorie-Zuordnungen
            keyword_to_category: Keyword-Kategorie-Mapping
            output_path: Zielpfad
        """
        # Berechne Per-Sheet-Häufigkeiten
        per_sheet_frequencies = self.calculate_per_sheet_frequencies(
            sheet_names, sheet_row_counts, category_assignments
        )
        
        # Berechne Gesamt-Häufigkeiten
        total_frequencies = self.calculate_category_frequencies(category_assignments)
        
        # Sammle Keywords (invertiere Mapping)
        keywords_per_category = self.collect_keywords_per_category(keyword_to_category)
        
        # Erstelle Workbook
        workbook = self.create_statistics_workbook(
            per_sheet_frequencies, total_frequencies, keywords_per_category
        )
        
        # Speichere
        self.save_statistics(workbook, output_path)

    def generate_pdf_statistics(self,
                                merged_results: List,  # List[MergedResult]
                                check_attributes: List,  # List[CheckAttribute]
                                keyword_to_category: Dict[str, str],
                                output_path: Path) -> None:
        """
        Erstellt Statistik-Datei für PDF-Analyseergebnisse.
        
        Args:
            merged_results: Liste von MergedResult-Objekten
            check_attributes: Prüfmerkmale
            keyword_to_category: Keyword-Kategorie-Mapping
            output_path: Zielpfad
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PDF-Statistiken"
        
        current_row = 1
        
        # Überschrift
        sheet.cell(row=current_row, column=1, value="PDF-Analyse Statistiken")
        current_row += 2
        
        # Gesamt-Statistiken
        sheet.cell(row=current_row, column=1, value="Gesamt-Übersicht")
        current_row += 1
        
        total_pdfs = len(merged_results)
        total_chunks = sum(r.chunk_count for r in merged_results)
        
        sheet.cell(row=current_row, column=1, value="Gesamt PDFs:")
        sheet.cell(row=current_row, column=2, value=total_pdfs)
        current_row += 1
        
        sheet.cell(row=current_row, column=1, value="Gesamt Chunks:")
        sheet.cell(row=current_row, column=2, value=total_chunks)
        current_row += 2
        
        # Keyword-Kategorie-Häufigkeiten
        sheet.cell(row=current_row, column=1, value="Keyword-Kategorien")
        current_row += 1
        
        sheet.cell(row=current_row, column=1, value="Kategorie")
        sheet.cell(row=current_row, column=2, value="Häufigkeit")
        sheet.cell(row=current_row, column=3, value="Keywords")
        current_row += 1
        
        # Zähle Kategorie-Häufigkeiten
        category_frequencies = {}
        for result in merged_results:
            if result.keyword_category:
                categories = [c.strip() for c in result.keyword_category.split(",")]
                for category in categories:
                    if category:
                        category_frequencies[category] = category_frequencies.get(category, 0) + 1
        
        # Sortiere nach Häufigkeit
        sorted_categories = sorted(category_frequencies.items(), key=lambda x: x[1], reverse=True)
        
        # Sammle Keywords pro Kategorie (invertiere Mapping)
        keywords_per_category = self.collect_keywords_per_category(keyword_to_category)
        
        # Schreibe Kategorie-Daten
        for category, frequency in sorted_categories:
            sheet.cell(row=current_row, column=1, value=category)
            sheet.cell(row=current_row, column=2, value=frequency)
            
            if category in keywords_per_category:
                keywords_str = ", ".join(sorted(keywords_per_category[category]))
                sheet.cell(row=current_row, column=3, value=keywords_str)
            
            current_row += 1
        
        current_row += 1
        
        # Custom Check Zusammenfassungen
        if check_attributes:
            sheet.cell(row=current_row, column=1, value="Prüfmerkmale-Zusammenfassung")
            current_row += 1
            
            sheet.cell(row=current_row, column=1, value="Prüfmerkmal")
            sheet.cell(row=current_row, column=2, value="Wert")
            sheet.cell(row=current_row, column=3, value="Häufigkeit")
            current_row += 1
            
            for attr in check_attributes:
                question = attr.question
                
                # Zähle Werte für dieses Prüfmerkmal
                value_counts = {}
                for result in merged_results:
                    value = result.custom_checks.get(question)
                    if value is not None:
                        # Konvertiere Boolean zu Ja/Nein
                        if attr.answer_type == "boolean":
                            if isinstance(value, bool):
                                display_value = "Ja" if value else "Nein"
                            else:
                                display_value = str(value)
                        else:
                            display_value = str(value)
                        
                        value_counts[display_value] = value_counts.get(display_value, 0) + 1
                
                # Schreibe Prüfmerkmal-Daten
                if value_counts:
                    sorted_values = sorted(value_counts.items(), key=lambda x: x[1], reverse=True)
                    for idx, (value, count) in enumerate(sorted_values):
                        if idx == 0:
                            sheet.cell(row=current_row, column=1, value=question)
                        sheet.cell(row=current_row, column=2, value=value)
                        sheet.cell(row=current_row, column=3, value=count)
                        current_row += 1
                else:
                    sheet.cell(row=current_row, column=1, value=question)
                    sheet.cell(row=current_row, column=2, value="(keine Daten)")
                    current_row += 1
        
        # Speichere
        self.save_statistics(workbook, output_path)
        logger.info(f"PDF-Statistiken erstellt: {output_path}")


    # ──────────────────────────────────────────────────────────────
    # Erweiterte Statistiken (Konfidenzintervalle, Intercoder)
    # ──────────────────────────────────────────────────────────────
    
    @staticmethod
    def calculate_confidence_interval(successes: int, total: int, 
                                      confidence_level: float = 0.95) -> tuple:
        """
        Berechnet Wilson-Score Konfidenzintervall für eine Häufigkeit.
        
        Args:
            successes: Anzahl Erfolge (z.B. 'Ja'-Antworten)
            total: Gesamtzahl
            confidence_level: Konfidenzniveau (default: 0.95)
            
        Returns:
            Tuple (lower, upper) als Dezimalzahlen
        """
        import math
        
        if total == 0:
            return (0.0, 0.0)
        
        p = successes / total
        z = 1.96 if confidence_level == 0.95 else 2.576 if confidence_level == 0.99 else 1.645
        
        denominator = 1 + z**2 / total
        center = (p + z**2 / (2 * total)) / denominator
        margin = z * math.sqrt((p * (1 - p) + z**2 / (4 * total)) / total) / denominator
        
        return (max(0, center - margin), min(1, center + margin))
    
    def calculate_confidence_intervals(self, results, check_attributes) -> List[dict]:
        """
        Berechnet Konfidenzintervalle für alle Prüfmerkmale.
        
        Args:
            results: Liste der Analyseergebnisse
            check_attributes: Prüfmerkmale
            
        Returns:
            Liste von dicts mit question, value, count, ci_lower, ci_upper
        """
        ci_results = []
        
        for attr in check_attributes:
            question = attr.question
            value_counts = {}
            
            for result in results:
                value = result.custom_checks.get(question)
                if value is not None:
                    if attr.answer_type == "boolean":
                        display_value = "Ja" if value else "Nein"
                    elif attr.answer_type == "multi_categorical" and isinstance(value, list):
                        for v in value:
                            display_value = str(v)
                            value_counts[display_value] = value_counts.get(display_value, 0) + 1
                        continue
                    else:
                        display_value = str(value)
                    value_counts[display_value] = value_counts.get(display_value, 0) + 1
            
            total = sum(value_counts.values())
            for value, count in value_counts.items():
                lower, upper = self.calculate_confidence_interval(count, total)
                ci_results.append({
                    "question": question,
                    "value": value,
                    "count": count,
                    "total": total,
                    "percentage": count / total * 100 if total > 0 else 0,
                    "ci_lower": lower * 100,
                    "ci_upper": upper * 100,
                    "ci_width": (upper - lower) * 100
                })
        
        return ci_results
    
    def calculate_intercoder_statistics(self, intercoder_result) -> dict:
        """
        Berechnet Intercoder-Statistiken für ein IntercoderResult.
        
        Args:
            intercoder_result: IntercoderResult-Objekt
            
        Returns:
            Dict mit agreement_rate, kappa_scores, overall_kappa
        """
        if not intercoder_result or not intercoder_result.agreements:
            return {
                "agreement_rate": 0.0,
                "total_items": 0,
                "agreed_items": 0,
                "kappa_scores": {},
                "overall_kappa": 0.0,
                "overall_interpretation": "unbekannt"
            }
        
        agreements = intercoder_result.agreements
        agreed = sum(1 for v in agreements.values() if v)
        total = len(agreements)
        
        return {
            "agreement_rate": agreed / total if total > 0 else 0.0,
            "total_items": total,
            "agreed_items": agreed,
            "kappa_scores": intercoder_result.kappa_scores,
            "overall_kappa": intercoder_result.overall_kappa,
            "overall_interpretation": intercoder_result.overall_interpretation
        }
    
    def add_confidence_interval_sheet(self, workbook: Workbook, results, check_attributes):
        """
        Fügt Konfidenzintervall-Sheet zum Workbook hinzu.
        
        Args:
            workbook: Das Workbook
            results: Analyseergebnisse
            check_attributes: Prüfmerkmale
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        sheet = workbook.create_sheet(title="Konfidenzintervalle")
        current_row = 1
        
        # Überschrift
        cell = sheet.cell(row=current_row, column=1, value="Konfidenzintervalle (95% Wilson-Score)")
        cell.font = Font(bold=True, size=14)
        current_row += 2
        
        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ["Prüfmerkmal", "Wert", "Häufigkeit", "Prozent", "CI unten", "CI oben", "CI Breite"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1
        
        # Daten
        ci_results = self.calculate_confidence_intervals(results, check_attributes)
        for ci in ci_results:
            sheet.cell(row=current_row, column=1, value=ci["question"])
            sheet.cell(row=current_row, column=2, value=ci["value"])
            sheet.cell(row=current_row, column=3, value=ci["count"])
            sheet.cell(row=current_row, column=4, value=f'{ci["percentage"]:.1f}%')
            sheet.cell(row=current_row, column=5, value=f'{ci["ci_lower"]:.1f}%')
            sheet.cell(row=current_row, column=6, value=f'{ci["ci_upper"]:.1f}%')
            sheet.cell(row=current_row, column=7, value=f'{ci["ci_width"]:.1f}%')
            current_row += 1
        
        logger.info(f"Konfidenzintervalle-Sheet hinzugefügt: {len(ci_results)} Einträge")
