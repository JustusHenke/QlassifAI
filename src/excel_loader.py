"""Excel Loader & Sheet Parser"""

from pathlib import Path
from typing import List, Optional
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from models import SheetInfo
from logging_config import get_logger

logger = get_logger("excel_loader")


class NoCompatibleSheetsError(Exception):
    """Fehler wenn keine kompatiblen Sheets gefunden werden"""
    pass


class ExcelLoader:
    """Lädt Excel-Dateien und identifiziert kompatible Sheets mit Textspalten"""
    
    # Gültige Spaltennamen für Textspalten (case-insensitive)
    VALID_TEXT_COLUMNS = ["text", "antwort", "answer", "textantwort"]
    
    def __init__(self, custom_text_column: Optional[str] = None):
        """
        Initialisiert ExcelLoader.
        
        Args:
            custom_text_column: Optionaler benutzerdefinierter Textspaltenname
        """
        self.custom_text_column = custom_text_column
    
    def load_workbook(self, file_path: Path) -> Workbook:
        """
        Lädt Excel-Datei mit openpyxl.
        
        Args:
            file_path: Pfad zur Excel-Datei
            
        Returns:
            Geladenes Workbook
            
        Raises:
            FileNotFoundError: Wenn Datei nicht existiert
            Exception: Bei anderen Ladefehlern
        """
        if not file_path.exists():
            error_msg = f"Excel-Datei nicht gefunden: {file_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        try:
            workbook = load_workbook(file_path, data_only=True)
            logger.info(f"Excel-Datei geladen: {file_path}")
            logger.info(f"Anzahl Sheets: {len(workbook.sheetnames)}")
            return workbook
        except KeyError as e:
            # Manche .xlsx-Dateien enthalten kaputte Drawing/Image-Referenzen.
            # openpyxl versucht diese zu lesen und bricht dann mit KeyError ab
            # (z.B. "xl/drawings/drawing1.xml"). In read_only lädt openpyxl i.d.R.
            # ohne die Drawing-Teile, was für unsere Text-Extraktion genügt.
            msg = str(e)
            if "xl/drawings/" in msg or "drawing" in msg.lower():
                logger.warning(
                    "Excel-Datei enthält kaputte Drawing-Referenzen (%s). "
                    "Versuche Fallback mit read_only=True ...",
                    msg,
                )
                try:
                    workbook = load_workbook(file_path, data_only=True, read_only=True)
                    logger.info(f"Excel-Datei geladen (read_only): {file_path}")
                    logger.info(f"Anzahl Sheets: {len(workbook.sheetnames)}")
                    return workbook
                except Exception as e2:
                    error_msg = (
                        f"Fehler beim Laden der Excel-Datei (Fallback read_only): {e2}. "
                        "Tipp: Datei in Excel öffnen und als neue .xlsx speichern (oder Bilder/Diagramme entfernen)."
                    )
                    logger.error(error_msg)
                    raise Exception(error_msg)

            error_msg = f"Fehler beim Laden der Excel-Datei: {e}"
            logger.error(error_msg)
            raise Exception(error_msg)
        except Exception as e:
            error_msg = f"Fehler beim Laden der Excel-Datei: {e}"
            logger.error(error_msg)
            raise Exception(error_msg)
    
    @staticmethod
    def _normalize_header_value(value) -> str:
        """
        Normalisiert Header-Zellwerte für robustes Matching.
        - macht lowercase
        - entfernt BOM / Zero-Width / NBSP
        - kollabiert Whitespace
        """
        if value is None:
            return ""
        try:
            s = str(value)
        except Exception:
            return ""

        # Entferne häufige unsichtbare Zeichen
        s = (
            s.replace("\ufeff", "")  # BOM
            .replace("\u200b", "")   # zero-width space
            .replace("\u00a0", " ")  # NBSP
        )
        s = " ".join(s.strip().lower().split())
        return s

    def locate_text_column(self, sheet: Worksheet, max_header_rows: int = 5) -> Optional[tuple]:
        """
        Findet Index der Textspalte in den ersten N Zeilen.
        
        Args:
            sheet: Worksheet zum Durchsuchen
            max_header_rows: Maximale Anzahl Zeilen zum Durchsuchen
            
        Returns:
            Tuple (column_index, header_row_index) oder None
        """
        # Erstelle Liste der zu suchenden Spaltennamen
        if self.custom_text_column:
            # Wenn benutzerdefinierter Name angegeben, nur diesen suchen
            valid_columns = [self._normalize_header_value(self.custom_text_column)]
            logger.info(f"Suche nach benutzerdefinierter Textspalte: '{self.custom_text_column}'")
        else:
            # Sonst Standard-Spaltennamen verwenden
            valid_columns = self.VALID_TEXT_COLUMNS

        valid_columns = [self._normalize_header_value(v) for v in valid_columns]
        
        for row_idx in range(1, min(max_header_rows + 1, sheet.max_row + 1)):
            row = sheet[row_idx]
            
            for col_idx, cell in enumerate(row, start=1):
                if cell.value:
                    cell_value_norm = self._normalize_header_value(cell.value)

                    # Exaktmatch oder "enthält"-Match (z.B. "Textantwort (offen)" / "Textantworten")
                    if (
                        cell_value_norm in valid_columns
                        or any(v and v in cell_value_norm for v in valid_columns)
                    ):
                        logger.info(
                            f"Textspalte '{cell.value}' gefunden in Spalte {col_idx}, "
                            f"Zeile {row_idx}"
                        )
                        return (col_idx, row_idx)
        
        return None
    
    def identify_data_rows(self, sheet: Worksheet, text_column_index: int, 
                          header_row_index: int) -> List[int]:
        """
        Identifiziert nicht-leere Datenzeilen unter der Kopfzeile.
        Ignoriert versteckte Zeilen (gefilterte Daten).
        
        Args:
            sheet: Worksheet
            text_column_index: Index der Textspalte
            header_row_index: Index der Kopfzeile
            
        Returns:
            Liste von Zeilen-Indizes mit nicht-leerem Text (nur sichtbare Zeilen)
        """
        data_rows = []
        
        # Starte nach der Kopfzeile
        for row_idx in range(header_row_index + 1, sheet.max_row + 1):
            # Prüfe ob Zeile versteckt ist (gefiltert)
            row_dimension = sheet.row_dimensions.get(row_idx)
            if row_dimension and row_dimension.hidden:
                logger.debug(f"Überspringe versteckte Zeile {row_idx}")
                continue
            
            cell = sheet.cell(row=row_idx, column=text_column_index)
            
            # Prüfe ob Zelle nicht-leeren Text enthält
            if cell.value and str(cell.value).strip():
                data_rows.append(row_idx)
        
        logger.info(f"{len(data_rows)} nicht-leere Datenzeilen gefunden (gefilterte Zeilen ignoriert)")
        return data_rows
    
    def find_compatible_sheets(self, workbook: Workbook) -> List[SheetInfo]:
        """
        Findet Sheets mit 'text', 'Antwort' oder 'answer' Spalten.
        
        Args:
            workbook: Zu durchsuchendes Workbook
            
        Returns:
            Liste von SheetInfo-Objekten für kompatible Sheets
            
        Raises:
            NoCompatibleSheetsError: Wenn keine kompatiblen Sheets gefunden
        """
        compatible_sheets = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logger.info(f"Analysiere Sheet: {sheet_name}")
            
            # Suche Textspalte
            result = self.locate_text_column(sheet)
            
            if result:
                text_column_index, header_row_index = result
                
                # Identifiziere Datenzeilen
                data_rows = self.identify_data_rows(
                    sheet, text_column_index, header_row_index
                )
                
                if data_rows:
                    sheet_info = SheetInfo(
                        name=sheet_name,
                        sheet=sheet,
                        text_column_index=text_column_index,
                        header_row_index=header_row_index,
                        data_rows=data_rows
                    )
                    compatible_sheets.append(sheet_info)
                    logger.info(
                        f"✓ Sheet '{sheet_name}' ist kompatibel "
                        f"({len(data_rows)} Zeilen)"
                    )
                else:
                    logger.warning(
                        f"Sheet '{sheet_name}' hat Textspalte aber keine Daten"
                    )
            else:
                logger.info(f"Sheet '{sheet_name}' hat keine Textspalte")
        
        if not compatible_sheets:
            error_msg = "Keine kompatiblen Sheets gefunden"
            logger.error(error_msg)
            raise NoCompatibleSheetsError(error_msg)
        
        logger.info(
            f"{len(compatible_sheets)} kompatible(s) Sheet(s) gefunden"
        )
        return compatible_sheets
    
    def load_and_analyze(self, file_path: Path) -> List[SheetInfo]:
        """
        Kombiniert Laden und Analyse in einer Methode.
        
        Args:
            file_path: Pfad zur Excel-Datei
            
        Returns:
            Liste von SheetInfo-Objekten
            
        Raises:
            FileNotFoundError: Wenn Datei nicht existiert
            NoCompatibleSheetsError: Wenn keine kompatiblen Sheets
            Exception: Bei anderen Fehlern
        """
        workbook = self.load_workbook(file_path)
        return self.find_compatible_sheets(workbook)
