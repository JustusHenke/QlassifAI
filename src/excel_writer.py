"""Excel Writer für Ergebnisse"""

from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from models import CheckAttribute, AnalysisResult, SheetInfo
from logging_config import get_logger

logger = get_logger("excel_writer")


class ExcelWriter:
    """Schreibt Analyseergebnisse in neue Excel-Datei"""
    
    # Theme-Konstanten
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    LOW_CONFIDENCE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    def __init__(self, confidence_threshold: float = 0.7):
        """
        Initialisiert ExcelWriter.
        
        Args:
            confidence_threshold: Schwellwert für farbliche Hervorhebung (default: 0.7)
        """
        self.confidence_threshold = confidence_threshold
    
    # ──────────────────────────────────────────────────────────────
    # Shared Helper Methods
    # ──────────────────────────────────────────────────────────────
    
    def _apply_header_style(self, cell) -> None:
        """Wendet einheitliches Header-Styling auf eine Zelle an"""
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.alignment = self.HEADER_ALIGNMENT
    
    def _format_custom_check_value(self, value: Any, attr: CheckAttribute) -> str:
        """Formatiert einen Custom-Check-Wert für die Anzeige in Excel"""
        if value is None:
            return "nicht kodiert"
        
        if attr.answer_type == "boolean":
            if isinstance(value, bool):
                return "Ja" if value else "Nein"
            elif str(value).lower() == "true":
                return "Ja"
            elif str(value).lower() == "false":
                return "Nein"
            else:
                return str(value)
        elif attr.answer_type == "multi_categorical":
            if isinstance(value, list):
                return ", ".join(str(v) for v in value)
            else:
                return str(value)
        else:  # categorical
            return str(value).replace("|", ", ")
    
    def _build_result_headers(self, check_attributes: List[CheckAttribute], 
                              include_reasoning: bool,
                              include_confidence: bool = False,
                              extra_headers: Optional[List[str]] = None) -> List[str]:
        """
        Erstellt die Spaltenüberschriften für Analyseergebnisse.
        """
        headers = ["Paraphrase", "Sentiment", "Sentiment_Begründung", "Keywords"]
        
        for attr in check_attributes:
            headers.append(attr.question)
            if include_reasoning:
                headers.append(f"{attr.question} (Begründung)")
        
        # Konfidenz-Spalten hinzufügen
        if include_confidence:
            for attr in check_attributes:
                headers.append(f"{attr.question} (Konfidenz)")
        
        headers.append("Keyword_Kategorie")
        
        if extra_headers:
            headers.extend(extra_headers)
        
        return headers
    
    def _write_check_attributes_to_row(self, sheet: Worksheet, row_idx: int,
                                        start_col: int, result,
                                        check_attributes: List[CheckAttribute],
                                        include_reasoning: bool) -> int:
        """Schreibt Custom-Check-Werte in eine Zeile"""
        col_offset = 0
        
        for attr in check_attributes:
            question = attr.question
            value = result.custom_checks.get(question)
            reason = result.custom_checks_reasons.get(question, "")
            
            display_value = self._format_custom_check_value(value, attr)
            sheet.cell(row=row_idx, column=start_col + col_offset, value=display_value)
            col_offset += 1
            
            if include_reasoning:
                sheet.cell(row=row_idx, column=start_col + col_offset, 
                          value=reason if reason else "")
                col_offset += 1
        
        return col_offset
    
    def _write_confidence_to_row(self, sheet: Worksheet, row_idx: int,
                                  start_col: int, result,
                                  check_attributes: List[CheckAttribute]) -> int:
        """Schreibt Konfidenz-Scores in eine Zeile mit farblicher Hervorhebung"""
        col_offset = 0
        
        for attr in check_attributes:
            question = attr.question
            score = result.confidence_scores.get(question)
            
            if score is not None:
                score_display = f"{score:.0%}"
                cell = sheet.cell(row=row_idx, column=start_col + col_offset, value=score_display)
                
                if score < self.confidence_threshold:
                    cell.fill = self.LOW_CONFIDENCE_FILL
            else:
                sheet.cell(row=row_idx, column=start_col + col_offset, value="-")
            
            col_offset += 1
        
        return col_offset
    
    def _write_stats_section(self, stats_sheet: Worksheet, current_row: int,
                             title: str, frequencies: Dict[str, int],
                             keywords_per_category: Dict[str, set]) -> int:
        """Schreibt eine Statistik-Sektion"""
        cell = stats_sheet.cell(row=current_row, column=1, value=title)
        cell.font = Font(bold=True, size=12)
        current_row += 1
        
        for col_idx, header in enumerate(["Kategorie", "Häufigkeit", "Keywords"], start=1):
            cell = stats_sheet.cell(row=current_row, column=col_idx, value=header)
            self._apply_header_style(cell)
        current_row += 1
        
        sorted_categories = sorted(frequencies.items(), key=lambda x: x[1], reverse=True)
        
        for category, frequency in sorted_categories:
            stats_sheet.cell(row=current_row, column=1, value=category)
            stats_sheet.cell(row=current_row, column=2, value=frequency)
            
            if category in keywords_per_category:
                keywords_str = ", ".join(sorted(keywords_per_category[category]))
                stats_sheet.cell(row=current_row, column=3, value=keywords_str)
            
            current_row += 1
        
        return current_row + 1
    
    def _write_check_attributes_stats(self, stats_sheet: Worksheet, current_row: int,
                                       check_attributes: List[CheckAttribute],
                                       all_results: List) -> int:
        """Schreibt Prüfmerkmal-Zusammenfassungen in das Statistik-Sheet"""
        if not check_attributes:
            return current_row
        
        cell = stats_sheet.cell(row=current_row, column=1, value="Prüfmerkmale-Zusammenfassung")
        cell.font = Font(bold=True)
        current_row += 1
        
        for col_idx, header in enumerate(["Prüfmerkmal", "Wert", "Häufigkeit"], start=1):
            cell = stats_sheet.cell(row=current_row, column=col_idx, value=header)
            self._apply_header_style(cell)
        current_row += 1
        
        for attr in check_attributes:
            question = attr.question
            value_counts = {}
            
            for result in all_results:
                value = result.custom_checks.get(question)
                if value is not None:
                    display_value = self._format_custom_check_value(value, attr)
                    
                    if attr.answer_type == "multi_categorical" and isinstance(value, list):
                        for v in value:
                            v_str = str(v)
                            value_counts[v_str] = value_counts.get(v_str, 0) + 1
                    else:
                        value_counts[display_value] = value_counts.get(display_value, 0) + 1
            
            if value_counts:
                sorted_values = sorted(value_counts.items(), key=lambda x: x[1], reverse=True)
                for idx, (value, count) in enumerate(sorted_values):
                    if idx == 0:
                        stats_sheet.cell(row=current_row, column=1, value=question)
                    stats_sheet.cell(row=current_row, column=2, value=value)
                    stats_sheet.cell(row=current_row, column=3, value=count)
                    current_row += 1
            else:
                stats_sheet.cell(row=current_row, column=1, value=question)
                stats_sheet.cell(row=current_row, column=2, value="(keine Daten)")
                current_row += 1
        
        return current_row
    
    def _write_confidence_stats(self, stats_sheet: Worksheet, current_row: int,
                                 check_attributes: List[CheckAttribute],
                                 all_results: List) -> int:
        """Schreibt Konfidenz-Statistiken in das Statistik-Sheet"""
        # Prüfe ob Konfidenz-Daten vorhanden sind
        has_confidence = any(
            hasattr(r, 'confidence_scores') and r.confidence_scores 
            for r in all_results
        )
        
        if not has_confidence:
            return current_row
        
        cell = stats_sheet.cell(row=current_row, column=1, value="Konfidenz-Statistiken")
        cell.font = Font(bold=True)
        current_row += 1
        
        for col_idx, header in enumerate(["Prüfmerkmal", "Ø Konfidenz", "Min", "Max", "Niedrig"], start=1):
            cell = stats_sheet.cell(row=current_row, column=col_idx, value=header)
            self._apply_header_style(cell)
        current_row += 1
        
        for attr in check_attributes:
            question = attr.question
            scores = [
                r.confidence_scores.get(question) 
                for r in all_results 
                if hasattr(r, 'confidence_scores') and r.confidence_scores.get(question) is not None
            ]
            
            if scores:
                avg = sum(scores) / len(scores)
                min_val = min(scores)
                max_val = max(scores)
                low_count = sum(1 for s in scores if s < self.confidence_threshold)
                
                stats_sheet.cell(row=current_row, column=1, value=question)
                stats_sheet.cell(row=current_row, column=2, value=f"{avg:.0%}")
                stats_sheet.cell(row=current_row, column=3, value=f"{min_val:.0%}")
                stats_sheet.cell(row=current_row, column=4, value=f"{max_val:.0%}")
                stats_sheet.cell(row=current_row, column=5, value=low_count)
                current_row += 1
            else:
                stats_sheet.cell(row=current_row, column=1, value=question)
                stats_sheet.cell(row=current_row, column=2, value="-")
                current_row += 1
        
        return current_row + 1
    
    def _collect_keywords_per_category(self, keyword_to_category: Dict[str, str]) -> Dict[str, set]:
        """Invertiert das Keyword->Kategorie Mapping"""
        keywords_per_category = {}
        for keyword, category in keyword_to_category.items():
            if category not in keywords_per_category:
                keywords_per_category[category] = set()
            keywords_per_category[category].add(keyword)
        return keywords_per_category
    
    def _check_has_confidence(self, all_results: List) -> bool:
        """Prüft ob Ergebnisse Konfidenz-Daten enthalten"""
        return any(
            hasattr(r, 'confidence_scores') and r.confidence_scores 
            for r in all_results
        )
    
    # ──────────────────────────────────────────────────────────────
    # Excel-Modus
    # ──────────────────────────────────────────────────────────────
    
    def create_new_workbook_with_results(self, 
                                        sheet_infos: List[SheetInfo],
                                        all_results: List[AnalysisResult],
                                        category_assignments: List[List[str]],
                                        check_attributes: List[CheckAttribute],
                                        keyword_to_category: Dict[str, str],
                                        output_path: Path,
                                        include_reasoning: bool = True,
                                        include_confidence: bool = False) -> None:
        """Erstellt Excel-Datei mit Analyseergebnissen und Statistiken"""
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)
        
        keywords_per_category = self._collect_keywords_per_category(keyword_to_category)
        has_confidence = include_confidence and self._check_has_confidence(all_results)
        
        result_idx = 0
        
        for sheet_info in sheet_infos:
            new_sheet = new_workbook.create_sheet(title=sheet_info.name)
            original_sheet = sheet_info.sheet
            col_count = original_sheet.max_column
            
            # Header kopieren
            for col_idx in range(1, col_count + 1):
                cell_value = original_sheet.cell(row=sheet_info.header_row_index, column=col_idx).value
                header_cell = new_sheet.cell(row=1, column=col_idx, value=cell_value)
                self._apply_header_style(header_cell)
            
            # Neue Spaltenüberschriften
            result_headers = self._build_result_headers(
                check_attributes, include_reasoning, include_confidence=has_confidence
            )
            start_col = col_count + 1
            
            for idx, col_name in enumerate(result_headers):
                header_cell = new_sheet.cell(row=1, column=start_col + idx, value=col_name)
                self._apply_header_style(header_cell)
            
            # Datenzeilen
            new_row_idx = 2
            for original_row_idx in sheet_info.data_rows:
                # Originaldaten kopieren
                for col_idx in range(1, col_count + 1):
                    cell_value = original_sheet.cell(row=original_row_idx, column=col_idx).value
                    new_sheet.cell(row=new_row_idx, column=col_idx, value=cell_value)
                
                result = all_results[result_idx]
                categories = category_assignments[result_idx]
                
                # Basis-Felder
                new_sheet.cell(row=new_row_idx, column=start_col, value=result.paraphrase)
                new_sheet.cell(row=new_row_idx, column=start_col + 1, value=result.sentiment)
                new_sheet.cell(row=new_row_idx, column=start_col + 2, value=result.sentiment_reason)
                new_sheet.cell(row=new_row_idx, column=start_col + 3, value=", ".join(result.keywords))
                
                # Custom Checks
                check_offset = self._write_check_attributes_to_row(
                    new_sheet, new_row_idx, start_col + 4, 
                    result, check_attributes, include_reasoning
                )
                
                # Konfidenz-Spalten
                conf_offset = 0
                if has_confidence:
                    conf_offset = self._write_confidence_to_row(
                        new_sheet, new_row_idx, start_col + 4 + check_offset,
                        result, check_attributes
                    )
                
                # Keyword_Kategorie
                new_sheet.cell(row=new_row_idx, column=start_col + 4 + check_offset + conf_offset, 
                              value=", ".join(categories))
                
                new_row_idx += 1
                result_idx += 1
            
            if new_sheet.max_row > 0 and new_sheet.max_column > 0:
                new_sheet.auto_filter.ref = new_sheet.dimensions
        
        # Statistiken-Sheet
        self._add_statistics_sheet(new_workbook, sheet_infos, all_results, 
                                   category_assignments, check_attributes, 
                                   keywords_per_category, has_confidence)
        
        return new_workbook
        # logger.info(f"Neue Workbook mit Statistiken erstellt: {output_path}")
    
    def _add_statistics_sheet(self, workbook: Workbook, sheet_infos: List[SheetInfo],
                               all_results: List, category_assignments: List[List[str]],
                               check_attributes: List[CheckAttribute],
                               keywords_per_category: Dict[str, set],
                               has_confidence: bool = False) -> None:
        """Fügt Statistiken-Sheet zum Workbook hinzu"""
        stats_sheet = workbook.create_sheet(title="Statistiken")
        current_row = 1
        
        cell = stats_sheet.cell(row=current_row, column=1, value="Kategorie-Statistiken")
        cell.font = Font(bold=True, size=14)
        current_row += 3
        
        # Per-Sheet-Statistiken
        result_idx = 0
        for sheet_info in sheet_infos:
            sheet_row_count = len(sheet_info.data_rows)
            sheet_assignments = category_assignments[result_idx:result_idx + sheet_row_count]
            
            frequencies = {}
            for categories in sheet_assignments:
                for category in categories:
                    if category:
                        frequencies[category] = frequencies.get(category, 0) + 1
            
            current_row = self._write_stats_section(
                stats_sheet, current_row, f"Sheet: {sheet_info.name}",
                frequencies, keywords_per_category
            )
            
            result_idx += sheet_row_count
        
        # Gesamt-Statistiken
        total_frequencies = {}
        for categories in category_assignments:
            for category in categories:
                if category:
                    total_frequencies[category] = total_frequencies.get(category, 0) + 1
        
        current_row = self._write_stats_section(
            stats_sheet, current_row, "Zusammen",
            total_frequencies, keywords_per_category
        )
        
        # Prüfmerkmal-Zusammenfassungen
        current_row = self._write_check_attributes_stats(
            stats_sheet, current_row, check_attributes, all_results
        )
        
        # Konfidenz-Statistiken
        if has_confidence:
            self._write_confidence_stats(
                stats_sheet, current_row, check_attributes, all_results
            )
    
    # ──────────────────────────────────────────────────────────────
    # PDF-Modus
    # ──────────────────────────────────────────────────────────────
    
    def create_pdf_results_workbook(self,
                                    merged_results: List,
                                    check_attributes: List[CheckAttribute],
                                    keyword_to_category: Dict[str, str],
                                    output_path: Path,
                                    include_reasoning: bool = True,
                                    include_confidence: bool = False) -> None:
        """Erstellt Excel-Datei mit PDF-Analyseergebnissen"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Analyseergebnisse"
        
        keywords_per_category = self._collect_keywords_per_category(keyword_to_category)
        sentiment_map = {-1: "negativ", 0: "gemischt", 1: "positiv"}
        
        headers = self._build_result_headers(
            check_attributes, include_reasoning, 
            include_confidence=include_confidence,
            extra_headers=["Chunk_Anzahl"]
        )
        headers.insert(0, "Dateiname")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(cell)
        
        for row_idx, result in enumerate(merged_results, start=2):
            col_idx = 1
            
            sheet.cell(row=row_idx, column=col_idx, value=result.filename); col_idx += 1
            sheet.cell(row=row_idx, column=col_idx, value=result.paraphrase); col_idx += 1
            sheet.cell(row=row_idx, column=col_idx, value=sentiment_map.get(result.sentiment, "gemischt")); col_idx += 1
            sheet.cell(row=row_idx, column=col_idx, value=result.sentiment_reason); col_idx += 1
            sheet.cell(row=row_idx, column=col_idx, value=", ".join(result.keywords) if result.keywords else ""); col_idx += 1
            
            check_offset = self._write_check_attributes_to_row(
                sheet, row_idx, col_idx, result, check_attributes, include_reasoning
            )
            col_idx += check_offset
            
            sheet.cell(row=row_idx, column=col_idx, value=result.keyword_category); col_idx += 1
            sheet.cell(row=row_idx, column=col_idx, value=result.chunk_count)
        
        if sheet.max_row > 0 and sheet.max_column > 0:
            sheet.auto_filter.ref = sheet.dimensions
        
        self._add_pdf_statistics_sheet(workbook, merged_results, check_attributes, keywords_per_category)
        
        workbook.save(output_path)
        logger.info(f"PDF-Ergebnisse mit Statistiken gespeichert: {output_path}")
        print(f"\n✓ PDF-Analysedatei mit Statistiken erstellt: {output_path}")
    
    def _add_pdf_statistics_sheet(self, workbook: Workbook, merged_results: List,
                                   check_attributes: List[CheckAttribute],
                                   keywords_per_category: Dict[str, set]) -> None:
        """Fügt PDF-Statistiken-Sheet zum Workbook hinzu"""
        stats_sheet = workbook.create_sheet(title="Statistiken")
        current_row = 1
        
        cell = stats_sheet.cell(row=current_row, column=1, value="PDF-Analyse Statistiken")
        cell.font = Font(bold=True, size=14)
        current_row += 2
        
        cell = stats_sheet.cell(row=current_row, column=1, value="Gesamt-Übersicht")
        cell.font = Font(bold=True)
        current_row += 1
        
        total_pdfs = len(merged_results)
        total_chunks = sum(r.chunk_count for r in merged_results)
        
        stats_sheet.cell(row=current_row, column=1, value="Gesamt PDFs:")
        stats_sheet.cell(row=current_row, column=2, value=total_pdfs); current_row += 1
        stats_sheet.cell(row=current_row, column=1, value="Gesamt Chunks:")
        stats_sheet.cell(row=current_row, column=2, value=total_chunks); current_row += 2
        
        category_frequencies = {}
        for result in merged_results:
            if result.keyword_category:
                for category in [c.strip() for c in result.keyword_category.split(",")]:
                    if category:
                        category_frequencies[category] = category_frequencies.get(category, 0) + 1
        
        current_row = self._write_stats_section(
            stats_sheet, current_row, "Keyword-Kategorien",
            category_frequencies, keywords_per_category
        )
        
        self._write_check_attributes_stats(stats_sheet, current_row, check_attributes, merged_results)

    # ──────────────────────────────────────────────────────────────
    # Intercoder-Sheet & Kappa-Sheet
    # ──────────────────────────────────────────────────────────────
    
    def create_intercoder_sheet(self, workbook: Workbook, 
                                 intercoder_result,
                                 check_attributes: List[CheckAttribute],
                                 sheet_name: str = "Intercoder") -> None:
        """
        Erstellt Intercoder-Vergleich-Sheet.
        
        Spalten: Text (gekürzt), Kodierung pro Modell, Übereinstimmung
        """
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Header erstellen
        headers = ["Text (gekürzt)"]
        for coder in intercoder_result.all_coder_results:
            headers.append(f"Kodierung ({coder.model_name})")
            headers.append(f"Konfidenz ({coder.model_name})")
        headers.append("Übereinstimmung")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(cell)
        
        # Für jedes Prüfmerkmal eine Zeile (vereinfacht: Text = Prüfmerkmal)
        row_idx = 2
        for attr in check_attributes:
            question = attr.question
            
            # Text (gekürzt)
            sheet.cell(row=row_idx, column=1, value=question[:50] + "..." if len(question) > 50 else question)
            
            col_idx = 2
            for coder in intercoder_result.all_coder_results:
                # Kodierung
                coding = coder.analysis_result.custom_checks.get(question, "nicht kodiert")
                sheet.cell(row=row_idx, column=col_idx, value=str(coding))
                col_idx += 1
                
                # Konfidenz
                conf_data = coder.confidence.get(question, {})
                score = conf_data.get("score")
                if score is not None:
                    cell = sheet.cell(row=row_idx, column=col_idx, value=f"{score:.0%}")
                    if score < self.confidence_threshold:
                        cell.fill = self.LOW_CONFIDENCE_FILL
                else:
                    cell = sheet.cell(row=row_idx, column=col_idx, value="-")
                col_idx += 1
            
            # Übereinstimmung
            agreement = intercoder_result.agreements.get(question, False)
            cell = sheet.cell(row=row_idx, column=col_idx, value="✓" if agreement else "✗")
            if not agreement:
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row_idx += 1
        
        # Autofilter aktivieren
        if sheet.max_row > 1:
            sheet.auto_filter.ref = f"A1:{sheet.dimensions.split(':')[1]}"
        
        logger.info(f"Intercoder-Sheet erstellt: {sheet_name}")
    
    def create_kappa_sheet(self, workbook: Workbook,
                           intercoder_result,
                           sheet_name: str = "Kappa-Statistik") -> None:
        """
        Erstellt Kappa-Statistik-Sheet.
        
        Spalten: Prüfmerkmal, Kappa, Interpretation, Konfidenzintervall
        """
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Header
        headers = ["Prüfmerkmal", "Kappa", "Interpretation", "Konfidenzintervall (±)", "N"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(cell)
        
        # Kappa-Ergebnisse pro Prüfmerkmal
        row_idx = 2
        for question, kappa_data in intercoder_result.kappa_scores.items():
            kappa = kappa_data.get("kappa", 0.0)
            interpretation = kappa_data.get("interpretation", "unbekannt")
            ci_width = kappa_data.get("ci_width", 0.0)
            n = kappa_data.get("n", 0)
            
            sheet.cell(row=row_idx, column=1, value=question)
            
            cell_kappa = sheet.cell(row=row_idx, column=2, value=f"{kappa:.3f}")
            # Farbliche Kodierung basierend auf Kappa
            if kappa >= 0.81:
                cell_kappa.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif kappa >= 0.61:
                cell_kappa.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif kappa >= 0.41:
                cell_kappa.fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
            else:
                cell_kappa.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
            
            sheet.cell(row=row_idx, column=3, value=interpretation)
            sheet.cell(row=row_idx, column=4, value=f"±{ci_width:.3f}" if ci_width > 0 else "-")
            sheet.cell(row=row_idx, column=5, value=n)
            
            row_idx += 1
        
        # Gesamt-Kappa
        row_idx += 1
        cell = sheet.cell(row=row_idx, column=1, value="GESAMT")
        cell.font = Font(bold=True, size=12)
        cell = sheet.cell(row=row_idx, column=2, value=f"{intercoder_result.overall_kappa:.3f}")
        cell.font = Font(bold=True)
        sheet.cell(row=row_idx, column=3, value=intercoder_result.overall_interpretation)
        sheet.cell(row=row_idx, column=3).font = Font(bold=True)
        
        logger.info(f"Kappa-Sheet erstellt: {sheet_name}")
